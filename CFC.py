#!/usr/bin python3
#!/usr/bin/env python3
# coding: utf-8
# Author: Helio Giroto
# Date: 17/02/2020

# ESTE SCRIPT É PARA GERAR/APPENDAR PLANILHA EXCEL DO MÊS REFERENTE AO PAGAMENTO DO C.F.C.

import openpyxl as op
import datetime, os

def mostraAjuda():
	print()
	print('\t-----------------')
	print('\t Tipo de Serviço')
	print('\t-----------------')
	print('\t  0- Ver opções')
	print('\t  1- EXAME MÉDICO')
	print('\t  2- EXAME PSICOTÉCNICO')
	print('\t  3- EXAME MÉDICO/PSICOTÉCNICO')
	print('\t  4- CURSO CFC PH')
	print('\t  5- CURSO RECICLAGEM')
	print('\t  6- CURSO RENOVAÇÃO')
	print('\t  7- AULA SIMULADOR')
	print('\t  8- ADMISSIONAL')
	print('\t  9- DEMISSIONAL')
	print('\t-----------------')
	print()


def mudaNome(nro):
	nro = nro.replace('1', 'EXAME MÉDICO')
	nro = nro.replace('2', 'EXAME PSICOTÉCNICO')
	nro = nro.replace('3', 'EXAME MÉDICO/PSICOTÉCNICO')
	nro = nro.replace('4', 'CURSO CFC PH')
	nro = nro.replace('5', 'CURSO RECICLAGEM')
	nro = nro.replace('6', 'CURSO RENOVAÇÃO')
	nro = nro.replace('7', 'AULA SIMULADOR')
	nro = nro.replace('8', 'ADMISSIONAL')
	nro = nro.replace('9', 'DEMISSIONAL')
	return nro


def salvaPlanilha(dadosServico):
	# Abre planilha para appendar:
	wb = op.load_workbook(planilha + '.xlsx')			
	ws = wb.active

	# appenda a lista digitada pelo usuário:
	ws.append(dadosServico)

	# nro da ultima linha da planilha:	
	ult_linha = ws.max_row
	ws.auto_filter.ref = 'A1:D' + str(ult_linha)


	# celulas laterais com totais de acordo à categoria do serviço:
	somaExMed  = '=SUMIF(C2:C'+str(ult_linha)+', "EXAME MÉDICO", D2:D'+str(ult_linha)+')'
	somaPsi    = '=SUMIF(C2:C'+str(ult_linha)+', "EXAME PSICOTÉCNICO", D2:D'+str(ult_linha)+')'
	somaMedPsi = '=SUMIF(C2:C'+str(ult_linha)+', "EXAME MÉDICO/PSICOTÉCNICO", D2:D'+str(ult_linha)+')'
	somaCfc    = '=SUMIF(C2:C'+str(ult_linha)+', "CURSO CFC PH", D2:D'+str(ult_linha)+')'
	somaRec    = '=SUMIF(C2:C'+str(ult_linha)+', "CURSO RECICLAGEM", D2:D'+str(ult_linha)+')'
	somaRen    = '=SUMIF(C2:C'+str(ult_linha)+', "CURSO RENOVAÇÃO", D2:D'+str(ult_linha)+')'
	somaSim    = '=SUMIF(C2:C'+str(ult_linha)+', "AULA SIMULADOR", D2:D'+str(ult_linha)+')'
	somaAdm    = '=SUMIF(C2:C'+str(ult_linha)+', "ADMISSIONAL", D2:D'+str(ult_linha)+')'
	somaDem    = '=SUMIF(C2:C'+str(ult_linha)+', "DEMISSIONAL", D2:D'+str(ult_linha)+')'
	celTotal   = '=SUM(D2:D'+str(ult_linha)+')'

	qtdeExMed  = '=COUNTIF(C2:C'+str(ult_linha)+', "EXAME MÉDICO")'
	qtdePsi    = '=COUNTIF(C2:C'+str(ult_linha)+', "EXAME PSICOTÉCNICO")'
	qtdeMedPsi = '=COUNTIF(C2:C'+str(ult_linha)+', "EXAME MÉDICO/PSICOTÉCNICO")'
	qtdeCfc    = '=COUNTIF(C2:C'+str(ult_linha)+', "CURSO CFC PH")'
	qtdeRec    = '=COUNTIF(C2:C'+str(ult_linha)+', "CURSO RECICLAGEM")'
	qtdeRen    = '=COUNTIF(C2:C'+str(ult_linha)+', "CURSO RENOVAÇÃO")'
	qtdeSim    = '=COUNTIF(C2:C'+str(ult_linha)+', "AULA SIMULADOR")'
	qtdeAdm    = '=COUNTIF(C2:C'+str(ult_linha)+', "ADMISSIONAL")'
	qtdeDem    = '=COUNTIF(C2:C'+str(ult_linha)+', "DEMISSIONAL")'
	qtdeTotal   = '=SUM(H2:H10)'

	# Poderia ser se não tivesse o recurso de sortear as colunas:
	# ws.cell(row=qtdeServicos+1, column=3).value = 'VALOR TOTAL: '
	# ws.cell(row=qtdeServicos+1, column=4).value = celTotal

	ws.cell(row=1, column=7).value = 'R$'
	ws.cell(row=1, column=8).value = 'Qtde'

	ws.cell(row=2, column=6).value = 'Exame Médico: '
	ws.cell(row=2, column=7).value = somaExMed
	ws.cell(row=2, column=8).value = qtdeExMed

	ws.cell(row=3, column=6).value = 'Exame Psicotécnico: '
	ws.cell(row=3, column=7).value = somaPsi
	ws.cell(row=3, column=8).value = qtdePsi

	ws.cell(row=4, column=6).value = 'Exame Médico/Psicotécnico: '
	ws.cell(row=4, column=7).value = somaMedPsi
	ws.cell(row=4, column=8).value = qtdeMedPsi

	ws.cell(row=5, column=6).value = 'Curso CFC PH: '
	ws.cell(row=5, column=7).value = somaCfc
	ws.cell(row=5, column=8).value = qtdeCfc

	ws.cell(row=6, column=6).value = 'Curso Reciclagem: '
	ws.cell(row=6, column=7).value = somaRec
	ws.cell(row=6, column=8).value = qtdeRec

	ws.cell(row=7, column=6).value = 'Curso Renovação: '
	ws.cell(row=7, column=7).value = somaRen
	ws.cell(row=7, column=8).value = qtdeRen

	ws.cell(row=8, column=6).value = 'Aula Simulador: '
	ws.cell(row=8, column=7).value = somaSim
	ws.cell(row=8, column=8).value = qtdeSim

	ws.cell(row=9, column=6).value = 'Admissional: '
	ws.cell(row=9, column=7).value = somaAdm
	ws.cell(row=9, column=8).value = qtdeAdm

	ws.cell(row=10, column=6).value = 'Demissional: '
	ws.cell(row=10, column=7).value = somaDem
	ws.cell(row=10, column=8).value = qtdeDem

	ws.cell(row=12, column=6).value = 'VALOR TOTAL: '
	ws.cell(row=12, column=7).value = celTotal
	ws.cell(row=12, column=8).value = qtdeTotal


	# salva:
	wb.save(planilha + '.xlsx')

	# PARA WINDOWS:
	# caminho = os.path.join("C:\\", "Users", "user", "Desktop", "python")
	# wb.save(caminho + '\\' + planilha + ".xlsx")

	print('\t\t*** Dados incluidos. ***')
	print()


def lancaDados():
	while True:
		print()
		print('Data do Serviço* : ', end='')
		try:		
			data = input().upper().strip(' ')
			data = data.replace('-','/')
			data = datetime.datetime.strptime(data, '%d/%m/%Y')
		except:
			if data == 'FIM':
				# salvaPlanilha(dadosServico)			# necessario ???	
				exit()
			else:
				print('\n***Erro na data. Repita digitação...***')
				lancaDados()


		print('Nome do Aluno....: ', end='')
		nome = input().upper()
		nome = nome.strip(' ')


		print('Tipo do Serviço..: ', end='')
		servico = input()
		servico = servico.strip(' ')
		if servico == '0':
			mostraAjuda()
			print('Tipo do Serviço..: ', end='')
			servico = input()


		nomeTipo = mudaNome(servico)

		print('Valor cobrado R$ : ', end='')
		valor = input()
		valor = valor.replace(',', '.').replace(' ', '')
		valor = float(valor)


		# junta dados inseridos numa lista:
		dadosServico = [data, nome, nomeTipo, valor]
		print(dadosServico)

		print()
		print("Confirma ?")
		print("----------") 
		print("\t(S)IM \n\t(N)ÃO ", end='')
		confirma = input()

		if confirma == "n" or confirma == "N":
			#data = "FIM"
			print('Registro cancelado... Redigite por favor:')
			print()		
		else:
			salvaPlanilha(dadosServico)
			# os.system('clear')
			# windows:
			# os.system('cls')
	# print()



print()	
print('*********************')
print('** CONTROLE DE CFC **')
print('*********************')
print()

print('Planilha a ser usada: ')
planilha = input().lower().strip()


# Se a planilha nao existe, cria uma nova:
if not os.path.exists(planilha + '.xlsx'):
	print('Esta planilha NÃO existe, deseja criar nova planilha com este nome?')
	print('(S)im / (N)ão : ', end='')
	criar = input().strip().lower()[:1]
	if criar == 'n':
		exit()
	else:
		from openpyxl import Workbook
		wb = Workbook()
		ws = wb.active
		cabecalho = ['DATA', 'NOME', 'SERVIÇO', 'VALOR']
		ws.append(cabecalho)
		wb.save(planilha + '.xlsx')
		print('Nova planilha criada: ' + planilha + '.xlsx')
		print()

print()
print('\t(- Digite FIM em "DATA" para terminar -)')
print('\t(- Digite 0 em "TIPO" para ver opções -)')
print()


lancaDados()

