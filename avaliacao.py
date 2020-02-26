#!/usr/bin python3
#!/usr/bin/env python3
# coding: utf-8
# Author: Helio Giroto
# Date: 24/02/2020

# ESTE SCRIPT ABRE PLANILHA E APPENDA DADOS. 
# Adiciona avaliações dos clientes.

# from openpyxl import Workbook 
import openpyxl as op
import datetime, os


def processa(novas_notas, planilha):

	# abre arquivo (w-book) de planilha:
	wb = op.load_workbook('avaliacao.xlsx')

	# Abre planilha (w-sheet) - aba: (podendo ser escolha do usuário)
	# ws = wb.get_sheet_by_name('AUTOESCOLA')
	ws = wb.get_sheet_by_name(planilha)
	
	# first_row = list(ws.rows)[0][0] # pega dados de coluna infinita....
	# print(first_row)

	# imprime o nro da última linha (do rodapé):
	print(ws.max_row)
	nro_ult_linha = ws.max_row
	

	# apaga a última linha: (linha antiga com totais de porcentagem)
	ws.delete_rows(nro_ult_linha)


	# manipulando lista de notas:
	sequenciaC 	= novas_notas[0]
	sequenciaN 	= novas_notas[1]
	nro_pesquisa 	= novas_notas[2]
	data 		= novas_notas[3]

	'''
	somaC = sum(novas_notas[0])	# erro = os itens da lista estão em str
	somaN = sum(novas_notas[1])	# erro = os itens da lista estão em str

	# porcentagem = soma / valor-max-notas * 100
	porcentagemC = ((somaC/32) * 100)
	porcentagemN = ((somaN/160) * 100)
	
	# ANTERIORMENTE - ABAIXO: erro por causa da divisão com 0
	for col in range(3, 32, 2):
		ws.cell(row=nro_ult_linha, column=col).value   = int(sequenciaC[(col-1)/2])
		ws.cell(row=nro_ult_linha, column=col+1).value = int(sequenciaN[(col-1)/2])
	'''


	# colocando itens das listas nas devidas celulas:
	for n in range(len(sequenciaC)):
		ws.cell(row=nro_ult_linha, column=n*2+1).value   = int(sequenciaC[n])
		ws.cell(row=nro_ult_linha, column=n*2+2).value   = int(sequenciaN[n])


	# Abaixo - porque não se pode dividir por zero (acima se o range fosse 1,32,2 - seria problema a conta con-1/2):
	ws.cell(row=nro_ult_linha, column=1).value   = sequenciaC[0]
	ws.cell(row=nro_ult_linha, column=2).value   = sequenciaN[0]

	ws.cell(row=nro_ult_linha, column=33).value   = nro_pesquisa
	ws.cell(row=nro_ult_linha, column=34).value   = data

	# para simplificar a sintaxe da concatenação:
	# todas formulas no openpyxl devem usar vírgula em vez de ponto e vírgula:
	N = str(nro_ult_linha)
	ws.cell(row=nro_ult_linha, column=35).value = "=((SUM(A"+N+",C"+N+",E"+N+",G"+N+",I"+N+",K"+N+",M"+N+",O"+N+",Q"+N+",S"+N+",U"+N+",W"+N+",Y"+N+",AA"+N+",AC"+N+",AE"+N+"))/32)*100"

	ws.cell(row=nro_ult_linha, column=36).value = "=((SUM(B"+N+",D"+N+",F"+N+",H"+N+",J"+N+",L"+N+",N"+N+",P"+N+",R"+N+",T"+N+",V"+N+",X"+N+",Z"+N+",AB"+N+",AD"+N+",AF"+N+"))/160)*100"



	# Appenda (na última linha) as novas notas (caso as listas estivessem em ordem conforme às colunas):
	# ws.append(novas_notas)


	# sort - ordem:
	ws.auto_filter.ref = 'A2:AJ' + str(nro_ult_linha)


	# O nro da ultima linha aumentou conforme foi appendando, por isso, se requer atualizar variável:
	nro_ult_linha = ws.max_row
	'''	
	L = ws.max_row
	
	for C in range(1, 32):
		ws.cell(row=L, column=C).value = '=(SUM(A3:A'+str(nro_ult_linha)+'))/(ROWS(A3:A'+str(nro_ult_linha)+')*2)*100'
	'''

	# =(SOMA(A3:A34))/(LINHAS(A3:A34)*2)*100
	tot_A = '=(SUM(A3:A'+str(nro_ult_linha)+'))/(ROWS(A3:A'+str(nro_ult_linha)+')*2)*100'
	tot_B = '=(SUM(B3:B'+str(nro_ult_linha)+'))/(ROWS(B3:B'+str(nro_ult_linha)+')*10)*100'
	tot_C = '=(SUM(C3:C'+str(nro_ult_linha)+'))/(ROWS(C3:C'+str(nro_ult_linha)+')*2)*100'
	tot_D = '=(SUM(D3:D'+str(nro_ult_linha)+'))/(ROWS(D3:D'+str(nro_ult_linha)+')*10)*100'
	tot_E = '=(SUM(E3:E'+str(nro_ult_linha)+'))/(ROWS(E3:E'+str(nro_ult_linha)+')*2)*100'
	tot_F = '=(SUM(F3:F'+str(nro_ult_linha)+'))/(ROWS(F3:F'+str(nro_ult_linha)+')*10)*100'
	tot_G = '=(SUM(G3:G'+str(nro_ult_linha)+'))/(ROWS(G3:G'+str(nro_ult_linha)+')*2)*100'
	tot_H = '=(SUM(H3:H'+str(nro_ult_linha)+'))/(ROWS(H3:H'+str(nro_ult_linha)+')*10)*100'
	tot_I = '=(SUM(I3:I'+str(nro_ult_linha)+'))/(ROWS(I3:I'+str(nro_ult_linha)+')*2)*100'
	tot_J = '=(SUM(J3:J'+str(nro_ult_linha)+'))/(ROWS(J3:J'+str(nro_ult_linha)+')*10)*100'
	tot_K = '=(SUM(K3:K'+str(nro_ult_linha)+'))/(ROWS(K3:K'+str(nro_ult_linha)+')*2)*100'
	tot_L = '=(SUM(L3:L'+str(nro_ult_linha)+'))/(ROWS(L3:L'+str(nro_ult_linha)+')*10)*100'
	tot_M = '=(SUM(M3:M'+str(nro_ult_linha)+'))/(ROWS(M3:M'+str(nro_ult_linha)+')*2)*100'
	tot_N = '=(SUM(N3:N'+str(nro_ult_linha)+'))/(ROWS(N3:N'+str(nro_ult_linha)+')*10)*100'
	tot_O = '=(SUM(O3:O'+str(nro_ult_linha)+'))/(ROWS(O3:O'+str(nro_ult_linha)+')*2)*100'
	tot_P = '=(SUM(P3:P'+str(nro_ult_linha)+'))/(ROWS(P3:P'+str(nro_ult_linha)+')*10)*100'
	tot_Q = '=(SUM(Q3:Q'+str(nro_ult_linha)+'))/(ROWS(Q3:Q'+str(nro_ult_linha)+')*2)*100'
	tot_R = '=(SUM(R3:R'+str(nro_ult_linha)+'))/(ROWS(R3:R'+str(nro_ult_linha)+')*10)*100'
	tot_S = '=(SUM(S3:S'+str(nro_ult_linha)+'))/(ROWS(S3:S'+str(nro_ult_linha)+')*2)*100'
	tot_T = '=(SUM(T3:T'+str(nro_ult_linha)+'))/(ROWS(T3:T'+str(nro_ult_linha)+')*10)*100'
	tot_U = '=(SUM(U3:U'+str(nro_ult_linha)+'))/(ROWS(U3:U'+str(nro_ult_linha)+')*2)*100'
	tot_V = '=(SUM(V3:V'+str(nro_ult_linha)+'))/(ROWS(V3:V'+str(nro_ult_linha)+')*10)*100'
	tot_W = '=(SUM(W3:W'+str(nro_ult_linha)+'))/(ROWS(W3:W'+str(nro_ult_linha)+')*2)*100'
	tot_X = '=(SUM(X3:X'+str(nro_ult_linha)+'))/(ROWS(X3:X'+str(nro_ult_linha)+')*10)*100'
	tot_Y = '=(SUM(Y3:Y'+str(nro_ult_linha)+'))/(ROWS(Y3:Y'+str(nro_ult_linha)+')*2)*100'
	tot_Z = '=(SUM(Z3:Z'+str(nro_ult_linha)+'))/(ROWS(Z3:Z'+str(nro_ult_linha)+')*10)*100'
	tot_AA = '=(SUM(AA3:AA'+str(nro_ult_linha)+'))/(ROWS(AA3:AA'+str(nro_ult_linha)+')*2)*100'
	tot_AB = '=(SUM(AB3:AB'+str(nro_ult_linha)+'))/(ROWS(AB3:AB'+str(nro_ult_linha)+')*10)*100'
	tot_AC = '=(SUM(AC3:AC'+str(nro_ult_linha)+'))/(ROWS(AC3:AC'+str(nro_ult_linha)+')*2)*100'
	tot_AD = '=(SUM(AD3:AD'+str(nro_ult_linha)+'))/(ROWS(AD3:AD'+str(nro_ult_linha)+')*10)*100'
	tot_AE = '=(SUM(AE3:AE'+str(nro_ult_linha)+'))/(ROWS(AE3:AE'+str(nro_ult_linha)+')*2)*100'
	tot_AF = '=(SUM(AF3:AF'+str(nro_ult_linha)+'))/(ROWS(AF3:AF'+str(nro_ult_linha)+')*10)*100'
	
	tot_AI = '=AVERAGE(AI3:AI'+str(nro_ult_linha)+')'
	tot_AJ = '=AVERAGE(AJ3:AJ'+str(nro_ult_linha)+')'


	# forma a linha de rodapé (uma lista) com totais:
	rodape = [tot_A, tot_B, tot_C, tot_D, tot_E, tot_F, tot_G, tot_H, tot_I, tot_J, tot_K, tot_L, tot_M, tot_N, tot_O, tot_P, tot_Q, tot_R, tot_S, tot_T, tot_U, tot_V, tot_W, tot_X, tot_Y, tot_Z, tot_AA, tot_AB, tot_AC, tot_AD, tot_AE, tot_AF]


	# appenda na planilha o rodapé com os totais:
	ws.append(rodape)


	# depois de appendar, é necessário colar na mesma linha de rodapé as médias das avaliações:
	ws.cell(row=nro_ult_linha+1, column=35).value = tot_AI
	ws.cell(row=nro_ult_linha+1, column=36).value = tot_AJ


	# para insertar uma linha em branco
	#ws.insert_rows(nro_ult_linha+1)


	# Salva planilha:
	wb.save('avaliacao.xlsx')

	print('\n**** Dados salvos na planilha. ****\n')


def lancaDadosDespachante():
	print()
	print('AVALIAÇÃO DE SERVIÇOS DE DESPACHANTE')
	print('------------------------------------')
	while True:
		print('Data: ', end='')
		try:		
			data = input().strip()
			data = data.replace('-','/')
			data = datetime.datetime.strptime(data, '%d/%m/%Y')
		except:
			data = data.upper()
			if data == 'FIM':
				exit()
			else:
				print('Data inválida, Favor redigitar...\n')
				lancaDadosAutoescola()
				
		print()
		print('Nro da Pesquisa: ', end='')
		nro_pesquisa = input().strip()
		print()

		print('**********************')
		print('**     CARINHAS     **')
		print('**      0 = :(      **')
		print('**      1 = :|      **')
		print('**      2 = :)      **')
		print('**********************')
		print()


		print('Digite a sequência de valores separados por espaço.')
		sequenciaC = input().strip().split(' ')[:16]
		print()
		print('\tLimpeza        : ', sequenciaC[0])		
		print('\tInfraestrutura : ', sequenciaC[1])
		print('\tInformações    : ', sequenciaC[2])
		print('\tHorário        : ', sequenciaC[3])
		print('\tPreço          : ', sequenciaC[4])
		print('\tCondições      : ', sequenciaC[5])
		print('\tAtendimento    : ', sequenciaC[6])
		print('\tOrientações    : ', sequenciaC[7])
		print('\tTratado        : ', sequenciaC[8])
		print('\tAgilidade      : ', sequenciaC[9])
		print('\tQualidade      : ', sequenciaC[10])
		print('\tPrazo          : ', sequenciaC[11])
		print('\tBrindes        : ', sequenciaC[12])
		print('\tSorteio        : ', sequenciaC[13])
		print('\tAvaliação      : ', sequenciaC[14])
		print('\tRecomenda      : ', sequenciaC[15])
		print()
		print('Confirma?\t(S)im\t(N)ão: ', end='')
		confirma = input()
		if confirma == 'n' or confirma == 'N':
			print()
			lancaDadosAutoescola()

		print()
		print('***********************')
		print('**       NOTAS       **')
		print('**                   **')
		print('**    (de 0 a 10)    **')
		print('**                   **')
		print('***********************')
		print()

		print('Digite a sequência de notas separadas por espaço.')
		sequenciaN = input().strip().split(' ')[:16]

		print()
		print('\tLimpeza        : ', sequenciaN[0])		
		print('\tInfraestrutura : ', sequenciaN[1])
		print('\tInformações    : ', sequenciaN[2])
		print('\tHorário        : ', sequenciaN[3])
		print('\tPreço          : ', sequenciaN[4])
		print('\tCondições      : ', sequenciaN[5])
		print('\tAtendimento    : ', sequenciaN[6])
		print('\tOrientações    : ', sequenciaN[7])
		print('\tTratado        : ', sequenciaN[8])
		print('\tAgilidade      : ', sequenciaN[9])
		print('\tQualidade      : ', sequenciaN[10])
		print('\tPrazo          : ', sequenciaN[11])
		print('\tBrindes        : ', sequenciaN[12])
		print('\tSorteio        : ', sequenciaN[13])
		print('\tAvaliação      : ', sequenciaN[14])
		print('\tRecomenda      : ', sequenciaN[15])
		print()
		print('Confirma?\t(S)im\t(N)ão: ', end='')
		confirma = input()
		if confirma == 'n' or confirma == 'N':
			print()
			lancaDadosAutoescola()

		print()

		os.system('clear')
		# caso o usuário lance mais de 16 itens é preciso só pegar os 16 e nada mais.

		# duas listas dentro de uma só:
		novas_notas = [sequenciaC, sequenciaN, nro_pesquisa, data]

		processa(novas_notas, 'DESPACHANTE')

def lancaDadosAutoescola():
	print()
	print('AVALIAÇÃO DE SERVIÇOS DE AUTOESCOLA')
	print('-----------------------------------')
	while True:
		print('Data: ', end='')
		try:		
			data = input().strip()
			data = data.replace('-','/')
			data = datetime.datetime.strptime(data, '%d/%m/%Y')
		except:
			data = data.upper()
			if data == 'FIM':
				exit()
			else:
				print('Data inválida, Favor redigitar...\n')
				lancaDadosAutoescola()
				
		print()
		print('Nro da Pesquisa: ', end='')
		nro_pesquisa = input().strip()
		print()

		print('**********************')
		print('**     CARINHAS     **')
		print('**      0 = :(      **')
		print('**      1 = :|      **')
		print('**      2 = :)      **')
		print('**********************')
		print()


		print('Digite a sequência de valores separados por espaço.')
		sequenciaC = input().strip().split(' ')[:16]
		print()
		print('\tLimpeza        : ', sequenciaC[0])		
		print('\tInfraestrutura : ', sequenciaC[1])
		print('\tFrota          : ', sequenciaC[2])
		print('\tInformações    : ', sequenciaC[3])
		print('\tHorário        : ', sequenciaC[4])
		print('\tPreço          : ', sequenciaC[5])
		print('\tCondições      : ', sequenciaC[6])
		print('\tAtendimento    : ', sequenciaC[7])
		print('\tOrientações    : ', sequenciaC[8])
		print('\tAgilidade      : ', sequenciaC[9])
		print('\tQualidade      : ', sequenciaC[10])
		print('\tPrazo          : ', sequenciaC[11])
		print('\tAulas Práticas : ', sequenciaC[12])
		print('\tSorteio        : ', sequenciaC[13])
		print('\tAvaliação      : ', sequenciaC[14])
		print('\tRecomenda      : ', sequenciaC[15])
		print()
		print('Confirma?\t(S)im\t(N)ão: ', end='')
		confirma = input()
		if confirma == 'n' or confirma == 'N':
			print()
			lancaDadosAutoescola()

		print()
		print('***********************')
		print('**       NOTAS       **')
		print('**                   **')
		print('**    (de 0 a 10)    **')
		print('**                   **')
		print('***********************')
		print()

		print('Digite a sequência de notas separadas por espaço.')
		sequenciaN = input().strip().split(' ')[:16]

		print()
		print('\tLimpeza        : ', sequenciaN[0])		
		print('\tInfraestrutura : ', sequenciaN[1])
		print('\tFrota          : ', sequenciaN[2])
		print('\tInformações    : ', sequenciaN[3])
		print('\tHorário        : ', sequenciaN[4])
		print('\tPreço          : ', sequenciaN[5])
		print('\tCondições      : ', sequenciaN[6])
		print('\tAtendimento    : ', sequenciaN[7])
		print('\tOrientações    : ', sequenciaN[8])
		print('\tAgilidade      : ', sequenciaN[9])
		print('\tQualidade      : ', sequenciaN[10])
		print('\tPrazo          : ', sequenciaN[11])
		print('\tAulas Práticas : ', sequenciaN[12])
		print('\tSorteio        : ', sequenciaN[13])
		print('\tAvaliação      : ', sequenciaN[14])
		print('\tRecomenda      : ', sequenciaN[15])
		print()
		print('Confirma?\t(S)im\t(N)ão: ', end='')
		confirma = input()
		if confirma == 'n' or confirma == 'N':
			print()
			lancaDadosAutoescola()

		print()

		os.system('clear')
		# caso o usuário lance mais de 16 itens é preciso só pegar os 16 e nada mais.

		# duas listas dentro de uma só:
		novas_notas = [sequenciaC, sequenciaN, nro_pesquisa, data]


		processa(novas_notas, 'AUTOESCOLA')



# só executa a função acima, se existir o arquivo avaliacao.xlxs:
if os.path.exists('avaliacao.xlsx'):
	print('Planilha existe.')
	print()
else:
	print('Planilha "avaliacao.xlsx" não existe.')


print()
print('Tipo de serviço que será avaliado:')
print('----------------------------------')
print('1 - Autoescola.')
print('2 - Despachante.')
tipo = input()
if tipo == '1':
	lancaDadosAutoescola()
elif tipo == '2':
	lancaDadosDespachante()
else:
	print('Opção não encontrada.')


